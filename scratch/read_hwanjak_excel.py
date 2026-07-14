import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"d:\OneDrive - 경상남도교육청\바탕 화면\진해고등학교\2026학년도\antigravity_folder\3학년_화법과작문_수행평가_양식.xlsx"

try:
    wb = openpyxl.load_workbook(excel_path)
    print("Sheets in workbook:", wb.sheetnames)
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"\nSheet: {name}, Max Row: {sheet.max_row}, Max Col: {sheet.max_column}")
        for r in range(1, min(15, sheet.max_row + 1)):
            row_vals = [sheet.cell(r, c).value for c in range(1, min(10, sheet.max_column + 1))]
            print(f"Row {r}: {row_vals}")
except Exception as e:
    print(f"Error reading Excel: {e}")
