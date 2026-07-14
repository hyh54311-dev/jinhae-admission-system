import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"d:\OneDrive - 경상남도교육청\바탕 화면\진해고등학교\2026학년도\antigravity_folder\청소년 SW동행프로젝트_학생_회원정보_양식.xlsx"

try:
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    print(f"Sheet Name: {sheet.title}")
    print(f"Max Row: {sheet.max_row}, Max Column: {sheet.max_column}")
    for row in range(1, sheet.max_row + 1):
        vals = [sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)]
        if any(vals):
            print(f"Row {row}: {vals}")
except Exception as e:
    print(f"Error: {e}")
