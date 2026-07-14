import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"d:\OneDrive - 경상남도교육청\바탕 화면\진해고등학교\2026학년도\antigravity_folder\3학년_화법과작문_수행평가_양식.xlsx"

if not os.path.exists(excel_path):
    print("Excel file not found at:", excel_path)
else:
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet_name = "수행평가_응답"
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"Sheet '{sheet_name}' total rows: {sheet.max_row}")
            print("Last 10 rows in Excel:")
            # Print last 10 rows
            start_row = max(1, sheet.max_row - 9)
            for r in range(start_row, sheet.max_row + 1):
                row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 10)]
                print(f"  Row {r}: {row_vals}")
        else:
            print(f"Sheet '{sheet_name}' not found in Excel. Available sheets: {wb.sheetnames}")
    except Exception as e:
        print("Error reading Excel:", e)
