import os
import openpyxl

def main():
    path = r"d:\OneDrive - 경상남도교육청\바탕 화면\진해고등학교\2026학년도\쿨메신져 다운로드 파일\2026. 창의적 체험활동 운영 계획(5.7.).xlsx"
    if not os.path.exists(path):
        print(f"File not found: {path}")
        return
        
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        
        with open('excel_inspection_report.txt', 'w', encoding='utf-8') as f:
            f.write(f"Sheet names: {wb.sheetnames}\n\n")
            
            for sheet_name in wb.sheetnames:
                f.write(f"\n--- Sheet: {sheet_name} ---\n")
                sheet = wb[sheet_name]
                # Print rows
                for r_idx in range(1, 100):
                    row_vals = []
                    for c_idx in range(1, 16):
                        cell_val = sheet.cell(row=r_idx, column=c_idx).value
                        row_vals.append(cell_val)
                    if any(row_vals is not None for row_vals in row_vals):
                        # Convert to string and clean up None
                        str_vals = [str(v) if v is not None else "" for v in row_vals]
                        f.write(f"Row {r_idx}: {str_vals}\n")
        print("Inspection completed. Report written to excel_inspection_report.txt")
    except Exception as e:
        print("Error reading excel:", e)

if __name__ == '__main__':
    main()
