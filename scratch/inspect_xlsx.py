import openpyxl

def main():
    filepath = r"d:\OneDrive - 경상남도교육청\바탕 화면\진해고등학교\2026학년도\antigravity_folder\3학년_화법과작문_수행평가_양식.xlsx"
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        for name in wb.sheetnames[:5]:
            sheet = wb[name]
            print(f"\nSheet {name} dimensions: {sheet.dimensions}")
            # Print first 5 rows
            for r_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if r_idx > 10:
                    break
                if any(row):
                    print(f"Row {r_idx}: {row[:10]}")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    main()
