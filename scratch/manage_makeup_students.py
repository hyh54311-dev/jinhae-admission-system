import os
import json
import openpyxl

EXCEL_PATH = r"d:\OneDrive - 경상남도교육청\바탕 화면\진해고등학교\2026학년도\antigravity_folder\3학년_화법과작문_수행평가_양식.xlsx"
DATA_PATH = r"d:\OneDrive - 경상남도교육청\바탕 화면\진해고등학교\2026학년도\antigravity_folder\scratch\makeup_data.json"
MD_PATH = r"d:\OneDrive - 경상남도교육청\바탕 화면\진해고등학교\2026학년도\antigravity_folder\수행평가_미응시자_명단.md"

def load_roster():
    """Load the complete student roster from Excel."""
    roster = {}
    if not os.path.exists(EXCEL_PATH):
        return roster
    
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if "명렬" in wb.sheetnames:
        sheet = wb["명렬"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            try:
                ban = int(row[0])
                num = int(row[1])
                name = str(row[2]).strip()
                s_id = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
                roster[(ban, num)] = {"name": name, "s_id": s_id}
            except Exception:
                continue
    return roster

def load_makeup_data():
    if os.path.exists(DATA_PATH):
        try:
            with open(DATA_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def save_makeup_data(data):
    os.makedirs(os.path.dirname(DATA_PATH), exist_ok=True)
    with open(DATA_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def add_makeup_student(ban, num):
    roster = load_roster()
    data = load_makeup_data()
    
    # Check if student already in the list
    for item in data:
        if item["ban"] == ban and item["num"] == num:
            return f"이미 등록된 학생입니다: {ban}반 {num}번 {item['name']}"
            
    student_info = roster.get((ban, num))
    if not student_info:
        # Fallback if not found in roster
        name = "Unknown"
        s_id = f"3{ban:02d}{num:02d}"
    else:
        name = student_info["name"]
        s_id = student_info["s_id"]
        
    new_student = {
        "ban": ban,
        "num": num,
        "name": name,
        "s_id": s_id,
        "status": "미응시" # "미응시", "응시완료"
    }
    
    data.append(new_student)
    # Sort by class and then by number
    data.sort(key=lambda x: (x["ban"], x["num"]))
    save_makeup_data(data)
    
    generate_markdown(data)
    return f"새로운 미응시 학생 추가 완료: {ban}반 {num}번 {name} (학번: {s_id})"

def generate_markdown(data):
    total_count = len(data)
    
    md_content = []
    md_content.append("# 3학년 화법과 작문 수행평가 미응시자 명단")
    md_content.append("\n> [!NOTE]")
    md_content.append("> **수행평가 미응시자 대상 추가 평가 안내**")
    md_content.append("> - **일시**: 다음 주 월요일 (6월 8일 예정) 7교시")
    md_content.append("> - **장소**: 화법과 작문실 (혹은 지정 장소)")
    md_content.append(f"> - **현재 대상 인원**: 총 **{total_count}명**")
    
    md_content.append("\n## 미응시자 명단\n")
    if not data:
        md_content.append("*등록된 미응시 학생이 없습니다.*")
    else:
        md_content.append("| 연번 | 학번 | 반 | 번호 | 이름 | 상태 | 비고 |")
        md_content.append("| :--- | :--- | :--- | :--- | :--- | :--- | :--- |")
        for idx, s in enumerate(data, 1):
            status_badge = "🔴 미응시" if s["status"] == "미응시" else "🟢 응시완료"
            md_content.append(f"| {idx} | {s['s_id']} | {s['ban']}반 | {s['num']}번 | **{s['name']}** | {status_badge} | |")
            
    md_content.append("\n## 반별 현황 요약\n")
    class_summary = {}
    for s in data:
        class_summary[s["ban"]] = class_summary.get(s["ban"], 0) + 1
        
    if not class_summary:
        md_content.append("*현황 요약 정보가 없습니다.*")
    else:
        md_content.append("| 반 | 미응시 인원 | 명단 |")
        md_content.append("| :--- | :--- | :--- |")
        for ban in sorted(class_summary.keys()):
            names = ", ".join([s["name"] for s in data if s["ban"] == ban])
            md_content.append(f"| {ban}반 | {class_summary[ban]}명 | {names} |")
            
    md_content.append("\n---\n*최종 업데이트: 2026-06-02*")
    
    with open(MD_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(md_content))

if __name__ == "__main__":
    import sys
    if len(sys.argv) >= 3:
        b = int(sys.argv[1])
        n = int(sys.argv[2])
        print(add_makeup_student(b, n))
    else:
        # Just regenerate if called with no args
        d = load_makeup_data()
        generate_markdown(d)
        print("Markdown file regenerated.")
