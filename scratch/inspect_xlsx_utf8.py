import openpyxl

def main():
    filepath = r"d:\OneDrive - 경상남도교육청\바탕 화면\진해고등학교\2026학년도\antigravity_folder\3학년_화법과작문_수행평가_양식.xlsx"
    outpath = r"d:\OneDrive - 경상남도교육청\바탕 화면\진해고등학교\2026학년도\antigravity_folder\scratch\inspect_xlsx_utf8.txt"
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        with open(outpath, "w", encoding="utf-8") as f:
            f.write(f"Sheets: {wb.sheetnames}\n")
            for name in wb.sheetnames:
                sheet = wb[name]
                f.write(f"\nSheet: {name}, dimensions: {sheet.dimensions}\n")
                for r_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if any(row):
                        f.write(f"Row {r_idx}: {list(row)}\n")
        print("Done! Check scratch/inspect_xlsx_utf8.txt")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    main()
