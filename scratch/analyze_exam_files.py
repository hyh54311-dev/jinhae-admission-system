import os
import sys

try:
    sys.stdout.reconfigure(encoding='utf-8')
except AttributeError:
    pass

output_path = r"d:\OneDrive - 경상남도교육청\바탕 화면\진해고등학교\2026학년도\antigravity_folder\scratch\exam_analysis_result.txt"
out_f = open(output_path, "w", encoding="utf-8")

def log(msg):
    out_f.write(msg + "\n")
    print(msg.encode('utf-8', errors='replace').decode('utf-8'))

def analyze_xlsx(file_path):
    log(f"\n=== Analyzing XLSX: {file_path} ===")
    if not os.path.exists(file_path):
        log("File does not exist.")
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            log(f"--- Sheet: {sheet_name} ---")
            sheet = wb[sheet_name]
            for r in range(1, sheet.max_row + 1):
                row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
                if any(row_vals):
                    while row_vals and row_vals[-1] is None:
                        row_vals.pop()
                    row_str = [str(x) if x is not None else "" for x in row_vals]
                    log(f"Row {r:03d}: " + " | ".join(row_str))
    except Exception as e:
        log(f"Error reading XLSX: {e}")

if __name__ == "__main__":
    path_lit_1 = r"D:\OneDrive - 경상남도교육청\바탕 화면\진해고등학교\2026학년도\수업\2026학년도 고사\2학년 1학기고사\중간고사\취합파일\2026학년도 1학기 1차고사 문학 나이스 문항정보표(수정 전).xlsx"
    path_lit_2 = r"D:\OneDrive - 경상남도교육청\바탕 화면\진해고등학교\2026학년도\수업\2026학년도 고사\2학년 1학기고사\중간고사\취합파일\문항 순서 수정\2026학년도 1학기 1차고사 문학 나이스 문항정보표.xlsx"
    
    analyze_xlsx(path_lit_1)
    analyze_xlsx(path_lit_2)
    
    out_f.close()
