import openpyxl

wb = openpyxl.load_workbook("scratch/submissions.xlsx")
print("Sheets in workbook:")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"Sheet '{name}': max_row={ws.max_row}, max_column={ws.max_column}")
    # Print the first 3 rows if max_row > 1
    if ws.max_row > 1:
        for r in range(1, min(ws.max_row + 1, 5)):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            print(f"  Row {r}: {row_vals[:10]}")
    print("-" * 30)
